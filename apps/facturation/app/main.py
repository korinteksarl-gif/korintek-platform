import os
import secrets
from fastapi import FastAPI, Request, HTTPException, Depends
from fastapi.responses import RedirectResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware

from . import auth, db

db.init_db()

app = FastAPI(title="KORINTEK — Facturation")

app.add_middleware(
    SessionMiddleware,
    secret_key=os.environ.get("SESSION_SECRET_KEY", secrets.token_hex(32)),
    same_site="lax",
    https_only=True,
)

templates = Jinja2Templates(directory=os.path.join(os.path.dirname(__file__), "templates"))


# ----------------------------------------------------------------------
# Dépendance : utilisateur connecté (redirige vers /login sinon)
# ----------------------------------------------------------------------

def get_current_user(request: Request):
    user = request.session.get("user")
    if not user:
        return None
    return user


def require_user(request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Non authentifié")
    return user


# ----------------------------------------------------------------------
# Auth routes
# ----------------------------------------------------------------------

@app.get("/login")
def login(request: Request):
    state = secrets.token_urlsafe(16)
    request.session["auth_state"] = state
    return RedirectResponse(auth.get_auth_url(state))


@app.get("/auth/callback")
def auth_callback(request: Request, code: str = None, state: str = None, error: str = None):
    if error:
        raise HTTPException(status_code=400, detail=f"Erreur Entra ID: {error}")
    if not code or state != request.session.get("auth_state"):
        raise HTTPException(status_code=400, detail="Requête d'authentification invalide")

    result = auth.acquire_token_by_code(code)
    claims = result.get("id_token_claims", {})
    email = (claims.get("preferred_username") or claims.get("email") or "").lower()
    name = claims.get("name", email)

    if not email:
        raise HTTPException(status_code=400, detail="Impossible de récupérer l'email du compte connecté")

    db.upsert_user(email, name)

    request.session["user"] = {
        "email": email,
        "name": name,
        "is_admin": auth.is_admin(email),
    }
    return RedirectResponse("/")


@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(auth.get_logout_url())


# ----------------------------------------------------------------------
# Page principale
# ----------------------------------------------------------------------

@app.get("/")
def index(request: Request):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    return templates.TemplateResponse("index.html", {"request": request, "user": user})


# ----------------------------------------------------------------------
# API — réglages société (lecture: tous ; écriture: admin uniquement)
# ----------------------------------------------------------------------

@app.get("/api/settings")
def api_get_settings(user=Depends(require_user)):
    return db.get_settings()


@app.post("/api/settings")
async def api_save_settings(request: Request, user=Depends(require_user)):
    if not user["is_admin"]:
        raise HTTPException(status_code=403, detail="Réservé à l'administrateur")
    data = await request.json()
    db.save_settings(data)
    return {"ok": True}


# ----------------------------------------------------------------------
# API — documents (proformas / factures)
# ----------------------------------------------------------------------

@app.get("/api/documents")
def api_list_documents(user=Depends(require_user)):
    return db.list_documents(user["email"], user["is_admin"])


@app.get("/api/documents/{doc_id}")
def api_get_document(doc_id: int, user=Depends(require_user)):
    doc = db.get_document(doc_id, user["email"], user["is_admin"])
    if not doc:
        raise HTTPException(status_code=404, detail="Document introuvable")
    return doc


@app.post("/api/documents")
async def api_create_document(request: Request, user=Depends(require_user)):
    doc = await request.json()
    doc_id = db.create_document(doc, created_by=user["email"])
    db.upsert_client(
        doc.get("client"), doc.get("clientTel"), doc.get("clientEmail"),
        doc.get("clientAdresse"), created_by=user["email"],
    )
    return {"ok": True, "id": doc_id}


# ----------------------------------------------------------------------
# API — base clients (autocomplétion)
# ----------------------------------------------------------------------

@app.get("/api/clients")
def api_search_clients(q: str = "", user=Depends(require_user)):
    if len(q.strip()) < 2:
        return []
    return db.search_clients(q.strip())


# ----------------------------------------------------------------------
# API — export comptable (Excel)
# ----------------------------------------------------------------------

@app.get("/api/documents/export.xlsx")
def api_export_documents(start: str, end: str, user=Depends(require_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    docs = db.get_documents_in_range(start, end, user["email"], user["is_admin"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Factures"

    headers = ["Type", "Numéro", "Date", "Client", "Téléphone", "HT (FCFA)",
               "TVA appliquée", "Retenue (%)", "Net à payer (FCFA)", "Créé par"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0E2226")

    for d in docs:
        ws.append([
            d["type"], d["num"], d["date"], d["client"], d["clientTel"],
            d["ht"], "Oui" if d["tvaOn"] else "Non", d["retenuePct"],
            d["total"], d["created_by"],
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"korintek_factures_{start}_a_{end}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ----------------------------------------------------------------------
# API — tableau de bord
# ----------------------------------------------------------------------

@app.get("/api/dashboard/stats")
def api_dashboard_stats(user=Depends(require_user)):
    return db.get_dashboard_stats(user["email"], user["is_admin"])


# ----------------------------------------------------------------------
# Qui suis-je (pour affichage cote client)
# ----------------------------------------------------------------------

@app.get("/api/me")
def api_me(user=Depends(require_user)):
    return user
